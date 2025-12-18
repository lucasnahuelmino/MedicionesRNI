from datetime import datetime
import numpy as np
import pandas as pd
import streamlit as st
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.drawing.image import Image as XLImage

from utils.time_utils import calcular_tiempo_total_por_archivo, format_timedelta_long


def render_resumen_general():
    # ------------------- RESUMEN GENERAL DE LOCALIDADES (con filtros previos) ------------------
    if "tabla_maestra" in st.session_state and not st.session_state["tabla_maestra"].empty:
        df = st.session_state["tabla_maestra"].copy()
        df["Resultado"] = pd.to_numeric(df["Resultado"], errors="coerce")

        # --- 🔍 FILTROS PREVIOS ---
        st.header("📊 Resumen general de mediciones")

        col1, col2, col3 = st.columns([1, 1, 1])
        with col1:
            ccte_sel = st.selectbox(
                "Filtrar CCTE",
                ["Todos"] + sorted(df["CCTE"].dropna().unique().tolist()),
                key="resumen_ccte"
            )
            if ccte_sel != "Todos":
                df = df[df["CCTE"] == ccte_sel]

        with col2:
            prov_sel = st.selectbox(
                "Filtrar Provincia",
                ["Todas"] + sorted(df["Provincia"].dropna().unique().tolist()),
                key="resumen_provincia"
            )
            if prov_sel != "Todas":
                df = df[df["Provincia"] == prov_sel]

        with col3:
            año_sel = "Todos"
            if "Fecha" in df.columns:
                df["Fecha"] = pd.to_datetime(df["Fecha"], dayfirst=True, errors="coerce")
                años_disp = sorted(df["Fecha"].dt.year.dropna().astype(int).unique().tolist(), reverse=True)
                if años_disp:
                    año_sel = st.selectbox(
                        "Filtrar Año",
                        ["Todos"] + [str(a) for a in años_disp],
                        key="resumen_año"
                    )
                    if año_sel != "Todos":
                        df = df[df["Fecha"].dt.year == int(año_sel)]

        # --- Procesamiento base ---
        if "Fecha" in df.columns:
            df["Fecha"] = pd.to_datetime(df["Fecha"], dayfirst=True, errors='coerce').dt.date
        if "Hora" in df.columns:
            df["Hora"] = pd.to_datetime(df["Hora"], errors='coerce').dt.time

        if "Fecha" in df.columns and "Hora" in df.columns:
            df["FechaHora"] = df.apply(
                lambda x: datetime.combine(x["Fecha"], x["Hora"]) if pd.notna(x["Fecha"]) and pd.notna(x["Hora"]) else pd.NaT,
                axis=1
            )
        else:
            df["FechaHora"] = pd.NaT

        # --- Resumen agrupado ---
        resumen_localidad = []
        for (ccte, prov, loc), g in df.groupby(["CCTE", "Provincia", "Localidad"]):
            inicio = g["FechaHora"].min() if "FechaHora" in g.columns else None
            fin = g["FechaHora"].max() if "FechaHora" in g.columns else None
            tiempo_total_localidad = calcular_tiempo_total_por_archivo(g)
            max_res = g["Resultado"].max() if pd.notna(g["Resultado"].max()) else None
            resumen_localidad.append({
                "CCTE": ccte,
                "Provincia": prov,
                "Localidad": loc,
                "Inicio": inicio,
                "Fin": fin,
                "Mediciones": len(g),
                "Tiempo mediciones": format_timedelta_long(tiempo_total_localidad),
                "Resultado Max (V/m)": max_res,
                "Resultado Max (%)": max_res**2 / 3770 / 0.20021 * 100 if max_res else None,
                "N° Expediente": ", ".join(sorted(g["Expediente"].dropna().unique().astype(str))),
                "Sonda utilizada": ", ".join(sorted(g["Sonda"].dropna().unique().astype(str))) if "Sonda" in g.columns else "N/A"
            })

        resumen_localidad_df = pd.DataFrame(resumen_localidad)

        st.dataframe(resumen_localidad_df)

        # --- Botón de exportación ---
        if not resumen_localidad_df.empty:
            if st.button("📥 Exportar resumen filtrado a Excel"):
                try:
                    ruta_excel = "resumen_localidades_filtrado.xlsx"
                    resumen_localidad_df.to_excel(ruta_excel, index=False)
                    wb = openpyxl.load_workbook(ruta_excel)
                    ws = wb.active

                    # Logo institucional
                    try:
                        logo_path = "assets/enacom_logo.png"
                        img = XLImage(logo_path)
                        img.width, img.height = 200, 70
                        ws.add_image(img, "A1")
                        ws.insert_rows(1, amount=5)
                    except Exception as e:
                        st.warning(f"No se pudo insertar logo: {e}")

                    for cell in ws[6]:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                    for row in ws.iter_rows(min_row=7, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                        for cell in row:
                            cell.alignment = Alignment(horizontal="center", vertical="center")

                    wb.save(ruta_excel)
                    st.success(f"Archivo '{ruta_excel}' generado con formato y logo.")
                    with open(ruta_excel, "rb") as f:
                        st.download_button(
                            label="⬇️ Descargar Excel filtrado",
                            data=f,
                            file_name=ruta_excel,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                except Exception as e:
                    st.error(f"Error exportando Excel: {e}")
